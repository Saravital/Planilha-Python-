import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Dados simulados da tabela de contas a receber
data = {
    "Cliente": ["Cliente A", "Cliente B", "Cliente C", "Cliente D"],
    "Nº do Título": [1234, 1235, 1236, 1237],
    "Data de Vencimento": ["10/09/2025", "12/09/2025", "15/09/2025", "05/09/2025"],
    "Valor (R$)": [850.00, 1200.00, 2500.00, 670.00],
    "Situação": ["A vencer", "A vencer", "Vencido", "A vencer"]
}

df = pd.DataFrame(data)

# Criando o arquivo Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Contas a Receber"

# Cabeçalho personalizado
empresa = "Minha Empresa LTDA"
usuario = "João da Silva"
data_geracao = datetime.now().strftime("%d/%m/%Y")

ws["A1"] = "Empresa:"
ws["B1"] = empresa
ws["A2"] = "Relatório gerado por:"
ws["B2"] = usuario
ws["A3"] = "Data de geração:"
ws["B3"] = data_geracao

# Espaço entre cabeçalho e tabela
start_row = 5

# Inserir os dados da tabela
for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)
        # Aplicar destaque vermelho se o valor >= 1000
        if r_idx > start_row and c_idx == 4 and isinstance(value, (int, float)) and value >= 1000:
            for i in range(1, 6):  # Colunas A até E
                ws.cell(row=r_idx, column=i).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

# Rodapé com totalizador
total = df["Valor (R$)"].sum()
ws[f"A{r_idx+2}"] = "TOTAL GERAL:"
ws[f"D{r_idx+2}"] = total

# Autoajuste das colunas
for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

# Salvar o arquivo na pasta do projeto Java
import os
output_path = os.path.join(r"C:\Users\MYHOME\projetos\IDE INTELLEJ - JAVA ESTUDO DIO\JAVA DIO\src", "Relatorio_Contas_a_Receber.xlsx")
wb.save(output_path)
print(f"Arquivo gerado com sucesso em: {output_path}")
